# Syntax Sugar (синтаксический сахар)
from docxtpl import DocxTemplate
# ternary if operator
# Jinjia

from openpyxl.reader.excel import load_workbook
#from docxtpl.reader
string =  ' '
context = {}

doc = DocxTemplate('template.docx')

wb = load_workbook(filename='personal.xlsx')
sheet = wb.active
rows = sheet.iter_rows(min_row=2, values_only=True)

for row in rows:
    num, fio, gender = row
    f_name = f'invitation_{num}.docx'
    context = {
        'dear': 'Уважаемый' if gender == 'м' else 'Уважаемая',
        'fio' : fio ,
        'number': num
    }
    doc.render(context)
    doc.save(f_name)





"""    if gender == 'ж':
        string = 'Уважаемая'
    else:
        string = 'Уважаемый'
    string += fio + '!'
    string += f'\n Ваш пригласительный билет №{num} на встречу ...'
    print(string)
    #string = 'Уважаемая' if gender == 'ж'
    print(string)"""

#print(list(row))
#sheet = wb['Сотрудники']
#print(sheet['B4'].value, sheet['C4'].value)


#temp = 25
#print('Тепло') if temp > 25 else print('Прохладно')

# a =map(int, s)
# s = ipaddress.split('.')
# a = [int(i) for i in ipaddress.split('.') if int(i) > 0]
# b = [i for i in range(2, 12) if i % 2 == 0]
#b = [i for i in range(2, 11, 2)]
#print(b)
# a =[ i * i for i in range(1,11)]


# print(list(a))
# lambda a,b: a+b
